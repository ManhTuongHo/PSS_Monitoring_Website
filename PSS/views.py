from django.shortcuts import render, redirect
from django.http import HttpResponse
from .forms import UserForm, UserProfileForm, HubForm, LoadForm, PanelForm
from .models import Load, LoadStatus, Panel, PanelStatus, Hub, PaymentStatus, PowerExchange
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .decorators import authenticated_user
from django.contrib.auth.models import User
from datetime import datetime, timedelta, date, time
from random import uniform, getrandbits
from django.http import HttpResponseRedirect, JsonResponse
import openpyxl
from dateutil.relativedelta import relativedelta
import stripe

# Create your views here.
stripe.api_key = "sk_test_51HURfzLvS1Gw3tPu74iB4jkelMo6q7nXamydYoUtfhHHcjjrg1gdDeFi8hTOQhBtP1sy1DrnTZN2v1ebxD5Pr2Le00ay4ZvjG2"

def PanelDataInitiate(panel_obj):
    now = datetime.now()
    now = now.replace(minute=0, second=0, microsecond=0)
    start_time = datetime(now.year - 1, 1, 1, 1, 0, 0)
    hour_count = int((now - start_time).total_seconds() / 3600)
    for consid_datetime in (start_time + timedelta(hours=n) for n in range(hour_count + 1)):
        ave_daily_generation = 0
        if consid_datetime.month == 12 or 1 or 2:
            ave_daily_generation = 4.4 * panel_obj.rated_power * 1.23
        elif consid_datetime.month == 3 or 4 or 5:
            ave_daily_generation = 4.4 * panel_obj.rated_power * 0.91
        elif consid_datetime.month == 6 or 7 or 8:
            ave_daily_generation = 4.4 * panel_obj.rated_power * 0.7
        elif consid_datetime.month == 9 or 10 or 11:
            ave_daily_generation = 4.4 * panel_obj.rated_power * 1.16

        ave_hour_generation = 0
        if (consid_datetime.hour < 6) or (consid_datetime.hour) > 19:
            ave_hour_generation = 0
        elif consid_datetime.hour == 6 or 19:
            ave_hour_generation = (ave_daily_generation / 14) * 0.2
        elif consid_datetime.hour == 7 or 18:
            ave_hour_generation = (ave_daily_generation / 14) * 0.5
        elif consid_datetime.hour == 8 or 17:
            ave_hour_generation = (ave_daily_generation / 14) * 0.7
        elif consid_datetime.hour == 9 or 16:
            ave_hour_generation = ave_daily_generation / 14
        elif consid_datetime.hour == 10 or 15:
            ave_hour_generation = (ave_daily_generation / 14) * 1.3
        elif consid_datetime.hour == 11 or 14:
            ave_hour_generation = (ave_daily_generation / 14) * 1.5
        elif consid_datetime.hour == 12 or 13:
            ave_hour_generation = (ave_daily_generation / 14) * 1.8

        hour_generation = (1 + uniform(-0.5, 0.5)) * ave_hour_generation
        consid_hour_panel_status = PanelStatus(panel=panel_obj, generation=hour_generation,
                                               date_time=consid_datetime)
        consid_hour_panel_status.save()


def LoadDataInitiate(load_obj):
    now = datetime.now()
    now = now.replace(minute=0, second=0, microsecond=0)
    start_time = datetime(now.year - 1, 1, 1, 1, 0, 0)
    hour_count = int((now - start_time).total_seconds() / 3600)
    for consid_datetime in (start_time + timedelta(hours=n) for n in range(hour_count + 1)):
        hour_consumption = 0

        if load_obj.load_name == "LED lighting":
            ave_hour_consumption = 0.1
            hour_consumption = (1 + uniform(-1.0, 0.5)) * ave_hour_consumption
        elif load_obj.load_name == "PC":
            ave_hour_consumption = 0.3
            hour_consumption = ave_hour_consumption*getrandbits(1)
        elif load_obj.load_name == "Fridge":
            ave_hour_consumption = 0.15
            hour_consumption = (1+uniform(-0.3,0.3))*ave_hour_consumption
        elif load_obj.load_name == "TV":
            ave_hour_consumption = 0.12
            hour_consumption = ave_hour_consumption * getrandbits(1)

        consid_hour_load_status = LoadStatus(load = load_obj, consumption = hour_consumption, date_time=consid_datetime)
        consid_hour_load_status.save()

def HubDataUpdate(hub_obj):
    now = datetime.now()
    now = now.replace(minute=0, second=0, microsecond=0)
    for load in hub_obj.load_set:
        n = 0
        while not load.loadstatus_set.filter(date_time = now - timedelta(hours = n)).exist():
            n += 1

        start_time = now - timedelta(hours = n-1)
        for consid_datetime in (start_time + timedelta(hours=m) for m in range(n)):
            hour_consumption = 0

            if load.load_name == "LED lighting":
                ave_hour_consumption = 0.1
                hour_consumption = (1 + uniform(-1.0, 0.5)) * ave_hour_consumption
            elif load.load_name == "PC":
                ave_hour_consumption = 0.3
                hour_consumption = ave_hour_consumption * getrandbits(1)
            elif load.load_name == "Fridge":
                ave_hour_consumption = 0.15
                hour_consumption = (1 + uniform(-0.3, 0.3)) * ave_hour_consumption
            elif load.load_name == "TV":
                ave_hour_consumption = 0.12
                hour_consumption = ave_hour_consumption * getrandbits(1)

            consid_hour_load_status = LoadStatus(load=load, consumption=hour_consumption, date_time=consid_datetime)
            consid_hour_load_status.save()

    for panel in hub_obj.panel_set:
        n = 0
        while not panel.panelstatus_set.filter(date_time = now - timedelta(hours = n)).exist():
            n += 1

        start_time = now - timedelta(hours = n-1)
        for consid_datetime in (start_time + timedelta(hours=m) for m in range(n)):
            ave_daily_generation = 0
            if consid_datetime.month == 12 or 1 or 2:
                ave_daily_generation = 4.4 * panel.rated_power * 1.23
            elif consid_datetime.month == 3 or 4 or 5:
                ave_daily_generation = 4.4 * panel.rated_power * 0.91
            elif consid_datetime.month == 6 or 7 or 8:
                ave_daily_generation = 4.4 * panel.rated_power * 0.7
            elif consid_datetime.month == 9 or 10 or 11:
                ave_daily_generation = 4.4 * panel.rated_power * 1.16

            ave_hour_generation = 0
            if (consid_datetime.hour < 6) or (consid_datetime.hour) > 19:
                ave_hour_generation = 0
            elif consid_datetime.hour == 6 or 19:
                ave_hour_generation = (ave_daily_generation / 14) * 0.2
            elif consid_datetime.hour == 7 or 18:
                ave_hour_generation = (ave_daily_generation / 14) * 0.5
            elif consid_datetime.hour == 8 or 17:
                ave_hour_generation = (ave_daily_generation / 14) * 0.7
            elif consid_datetime.hour == 9 or 16:
                ave_hour_generation = ave_daily_generation / 14
            elif consid_datetime.hour == 10 or 15:
                ave_hour_generation = (ave_daily_generation / 14) * 1.3
            elif consid_datetime.hour == 11 or 14:
                ave_hour_generation = (ave_daily_generation / 14) * 1.5
            elif consid_datetime.hour == 12 or 13:
                ave_hour_generation = (ave_daily_generation / 14) * 1.8

            hour_generation = (1 + uniform(-0.5, 0.5)) * ave_hour_generation
            consid_hour_panel_status = PanelStatus(panel=panel, generation=hour_generation, date_time=consid_datetime)
            consid_hour_panel_status.save()


def daily_consumption_fnc(load, year, month, day):
    daily_consumption = 0
    for load_status in load.loadstatus_set.all():
        if (load_status.date_time.year == year) and (load_status.date_time.month == month) and (load_status.date_time.day == day):
            daily_consumption += load_status.consumption
    return round(daily_consumption, 2)


def monthly_consumption_fnc(load, year, month):
    monthly_consumption = 0
    for load_status in load.loadstatus_set.all():
        if (load_status.date_time.year == year) and (load_status.date_time.month == month):
            monthly_consumption += load_status.consumption
    return round(monthly_consumption, 2)


def yearly_consumption_fnc(load, year):
    yearly_consumption = 0
    for load_status in load.loadstatus_set.all():
        if load_status.date_time.year == year:
            yearly_consumption += load_status.consumption
    return round(yearly_consumption, 2)


def daily_generation_fnc(panel, year, month, day):
    daily_generation = 0
    for panel_status in panel.panelstatus_set.all():
        if (panel_status.date_time.year == year) and (panel_status.date_time.month == month) and (panel_status.date_time.day == day):
            daily_generation += panel_status.generation
    return round(daily_generation,2)


def monthly_generation_fnc(panel, year, month):
    monthly_generation = 0
    for panel_status in panel.panelstatus_set.all():
        if (panel_status.date_time.year == year) and (panel_status.date_time.month == month):
            monthly_generation += panel_status.generation
    return round(monthly_generation, 2)


def yearly_generation_fnc(panel, year):
    yearly_generation = 0
    for panel_status in panel.pvstatus_set.all():
        if panel_status.date_time.year == year:
            yearly_generation += panel_status.generation
    return round(yearly_generation, 2)


def getPreviousYearMonth(year, month):
    if month == 1:
        previous_month = 12
        previous_year = year-1
    else:
        previous_month = month-1
        previous_year = year
    return previous_year, previous_month


def getNextYearMonth(year, month):
    if month == 12:
        next_month = 1
        next_year = year+1
    else:
        next_month = month+1
        next_year = year
    return next_year, next_month


def powerExchangeUpdate(month, year):
    previous_year, previous_month = getPreviousYearMonth(year, month)

    for user in User.objects.all():
        if (((user.userprofile.start_month <= month) and (user.userprofile.start_year == year)) or (user.userprofile.start_year < year)) \
                and (user.powerexchange_set.filter(month=month, year=year).count() == 0):
            user_monthly_consumption = 0
            user_monthly_generation = 0
            for hub in user.hub_set.all():
                for load in hub.load_set.all():
                    user_monthly_consumption += monthly_consumption_fnc(load, year, month)

                for panel in hub.panel_set.all():
                    user_monthly_generation += monthly_generation_fnc(panel, year, month)

            if user.powerexchange_set.filter(month=previous_month, year=previous_year).count() != 0:
                pre_exchange = user.powerexchange_set.get(month=previous_month, year=previous_year)
                user_monthly_generation = user_monthly_generation + pre_exchange.contrib - pre_exchange.contrib_amount_system
            else:
                user_monthly_generation += 0

            if user_monthly_consumption > user_monthly_generation:
                exchange = PowerExchange(user = user, contrib = 0, consume = (user_monthly_consumption - user_monthly_generation), month=month, year=year)
                exchange.save()
            elif user_monthly_consumption < user_monthly_generation:
                exchange = PowerExchange(user = user, contrib = (user_monthly_generation - user_monthly_consumption), consume = 0, month=month, year=year)
                exchange.save()
            else:
                exchange = PowerExchange(user=user, contrib=0,consume=0, month=month, year=year)
                exchange.save()

    total_generation = 0
    total_install_cost = 0
    for user in User.objects.all():
        if ((user.userprofile.start_month <= month) and (user.userprofile.start_year == year)) or (user.userprofile.start_year < year):
            for hub in user.hub_set.all():
                for panel in hub.panel_set.all():
                    total_generation += monthly_generation_fnc(panel, year, month)
                    total_install_cost += panel.install_cost

    energy_rate_system = total_install_cost/(total_generation*300)

    total_over_generation = 0
    total_over_consumption = 0

    for user in User.objects.all():
        if ((user.userprofile.start_month <= month) and (user.userprofile.start_year == year)) or (user.userprofile.start_year < year):
            consid_exchange = user.powerexchange_set.get(year=year,month=month)
            if consid_exchange.contrib > 0:
                total_over_generation += consid_exchange.contrib
            elif consid_exchange.consume > 0:
                total_over_consumption += consid_exchange.consume

    if total_over_generation > total_over_consumption:
        for user in User.objects.all():
            if ((user.userprofile.start_month <= month) and (user.userprofile.start_year == year)) or (user.userprofile.start_year < year):
                consid_exchange = user.powerexchange_set.get(year=year, month=month)
                if consid_exchange.contrib > 0:
                    consid_exchange.contrib_amount_system = (consid_exchange.contrib/total_over_generation)*total_over_consumption
                    consid_exchange.consume_amount_system = 0
                    consid_exchange.energy_rate_system = energy_rate_system
                    consid_exchange.save()
                elif consid_exchange.consume > 0:
                    consid_exchange.contrib_amount_system = 0
                    consid_exchange.consume_amount_system = consid_exchange.consume
                    consid_exchange.energy_rate_system = energy_rate_system
                    consid_exchange.save()

    elif total_over_generation < total_over_consumption:
        for user in User.objects.all():
            if ((user.userprofile.start_month <= month) and (user.userprofile.start_year == year)) or (user.userprofile.start_year < year):
                consid_exchange = user.powerexchange_set.get(year=year, month=month)
                if consid_exchange.contrib > 0:
                    consid_exchange.contrib_amount_system = consid_exchange.contrib
                    consid_exchange.consume_amount_system = 0
                    consid_exchange.energy_rate_system = energy_rate_system
                    consid_exchange.save()
                elif consid_exchange.consume > 0:
                    consid_exchange.contrib_amount_system = 0
                    consid_exchange.consume_amount_system = (consid_exchange.consume/total_over_consumption)*total_over_generation
                    consid_exchange.energy_rate_system = energy_rate_system
                    consid_exchange.save()


def paymentCalculate():
    default_rate = 0.3
    for user in User.objects.all():
        for exchange in user.powerexchange_set.all():
            if user.paymentstatus_set.filter(month = exchange.month, year = exchange.year).count() == 0:
                if (exchange.contrib_amount_system > 0) and (exchange.consume == 0) and (exchange.consume_amount_system == 0):
                    payment = PaymentStatus(user=user, month=exchange.month, year=exchange.year, earn_or_pay = True, amount = round(exchange.contrib_amount_system*exchange.energy_rate_system, 2), pay_centre = 0)
                    payment.save()
                elif (exchange.contrib_amount_system == 0) and (exchange.consume == 0) and (exchange.consume_amount_system == 0):
                    payment = PaymentStatus(user=user, month=exchange.month, year=exchange.year, earn_or_pay = True, amount = 0, pay_centre = 0)
                    payment.save()
                elif (exchange.contrib_amount_system == 0) and (exchange.consume > 0) and (exchange.consume_amount_system == exchange.consume):
                    payment = PaymentStatus(user=user, month=exchange.month, year=exchange.year, earn_or_pay = False, amount = round(exchange.consume_amount_system*exchange.energy_rate_system,2), pay_centre = 0)
                    payment.save()
                elif (exchange.contrib_amount_system == 0) and (exchange.consume > 0) and (exchange.consume_amount_system == 0):
                    payment = PaymentStatus(user=user, month=exchange.month, year=exchange.year, earn_or_pay = False, amount = 0, pay_centre = round(exchange.consume*default_rate,2))
                    payment.save()
                elif (exchange.contrib_amount_system == 0) and (exchange.consume > 0) and (exchange.consume_amount_system < exchange.consume):
                    payment = PaymentStatus(user=user, month=exchange.month, year=exchange.year, earn_or_pay = False, amount = round(exchange.consume_amount_system*exchange.energy_rate_system,2),
                                            pay_centre = round((exchange.consume - exchange.consume_amount_system)*default_rate,2))
                    payment.save()


@authenticated_user
def Login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('PSS:home')
        else:
            messages.info(request, 'Username or Password is incorrect')
    context = {}
    return render(request, 'PSS/login.html', context)


@authenticated_user
def signup(request):
    user_form = UserForm()
    profile_form = UserProfileForm()

    if request.method == 'POST':
        user_form = UserForm(request.POST)
        profile_form = UserProfileForm(request.POST)
        if user_form.is_valid() and profile_form.is_valid():
            user = user_form.save()
            profile = profile_form.save(commit=False)
            profile.user = user
            profile.save()

            username = user_form.cleaned_data.get('username')
            password = user_form.cleaned_data.get('password1')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('PSS:home')

    context = {'form': user_form, 'profile_form': profile_form}
    return render(request, 'PSS/signup.html', context)


def Logout(request):
    logout(request)
    return redirect('PSS:login')


@login_required(login_url='PSS:login')
def homepage(request):
    hub_form = HubForm()
    load_form = LoadForm()
    panel_form = PanelForm()
    current_user = User.objects.get(username=request.user.username)
    today = date.today()
    current_year,current_month = getPreviousYearMonth(today.year, today.month)
    while (((current_user.userprofile.start_month <= current_month) and (current_user.userprofile.start_year==current_year)) or (current_user.userprofile.start_year < current_year))\
            and (current_user.powerexchange_set.filter(month = current_month, year=current_year).count() == 0):
        current_year, current_month = getPreviousYearMonth(current_year, current_month)
    current_year, current_month = getNextYearMonth(current_year, current_month)
    last_year, last_month = getPreviousYearMonth(today.year, today.month)
    while (current_year < last_year) or ((current_month <= last_month) and (current_year == last_year)):
        powerExchangeUpdate(current_month, current_year)
        current_year, current_month = getNextYearMonth(current_year, current_month)
    paymentCalculate()
    last_exchange = current_user.powerexchange_set.get(year = last_year, month= last_month)
    total_excessive_generation = round((last_exchange.contrib - last_exchange.contrib_amount_system), 2)
    exchange_array = []
    payment_array = []
    total_fee = 0
    for payment in current_user.paymentstatus_set.all():
        if ((payment.earn_or_pay == False) or (payment.pay_centre > 0)) and (payment.status == False):
            exchange = current_user.powerexchange_set.get(month = payment.month, year = payment.year)
            exchange_array.append(exchange)
            payment_array.append(payment)
            total_fee += (payment.amount + payment.pay_centre)
    total_fee = int(round(total_fee))
    context = {'hub_form': hub_form, 'load_form': load_form, 'panel_form': panel_form, 'user': current_user, 'payment_array':payment_array, 'exchange_array': exchange_array, 'excessive_generation':total_excessive_generation, 'total_fee': total_fee}
    if request.method == 'POST':
        hub_form = HubForm(request.POST)
        load_form = LoadForm(request.POST)
        panel_form = PanelForm(request.POST)

        if hub_form.is_valid():
            hub = hub_form.save(commit=False)
            hub.owner = current_user
            hub.save()

        if load_form.is_valid():
            load = load_form.save(commit=False)
            load.hub = current_user.hub_set.get(hub_name=request.POST.get('load_hub_name'))
            load.save()
            LoadDataInitiate(load)

        if panel_form.is_valid():
            panel = panel_form.save(commit=False)
            panel.hub = current_user.hub_set.get(hub_name=request.POST.get('panel_hub_name'))
            panel.save()
            PanelDataInitiate(panel)

    return render(request, 'PSS/homepage.html', context)

def hubDetail(request, hub_name):
    load_form = LoadForm()
    panel_form = PanelForm()
    current_user = User.objects.get(username = request.user.username)
    current_hub = current_user.hub_set.get(hub_name = hub_name)
    now = datetime.now()
    max_date = date.today()
    min_date = date(max_date.year - 1, 1, 1)
    current_year = now.year
    context = {'load_form': load_form, 'panel_form': panel_form, 'hub': current_hub, 'years': range(current_year,current_year-2,-1), 'months': range(1,13), 'min_date': min_date, 'max_date': max_date}
    if request.method == 'POST':
        load_form = LoadForm(request.POST)
        panel_form = PanelForm(request.POST)

        if load_form.is_valid():
            load = load_form.save(commit=False)
            load.hub = current_hub
            load.save()
            LoadDataInitiate(load)

        if panel_form.is_valid():
            panel = panel_form.save(commit=False)
            panel.hub = current_hub
            panel.save()
            LoadDataInitiate(panel)

    return render(request, 'PSS/hub_detail.html', context)


def loadDelete(request, hub_name, load_name):
    current_user = User.objects.get(username = request.user.username)
    current_hub = current_user.hub_set.get(hub_name = hub_name)
    current_load = current_hub.load_set.get(load_name = load_name)
    current_load.delete()

    return redirect('PSS:hub-detail', hub_name = hub_name)


def panelDelete(request, hub_name, panel_name):
    current_user = User.objects.get(username = request.user.username)
    current_hub = current_user.hub_set.get(hub_name = hub_name)
    current_panel = current_hub.panel_set.get(panel_name = panel_name)
    current_panel.delete()

    return redirect('PSS:hub-detail',hub_name = hub_name)


def hubYearData(request, hub_name):
    yearData = []
    consumption_dict = {}
    generation_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    if request.method == 'POST':
        year = request.POST.get('year')
        year = int(year)
        for month in range(1,13):
            hub_monthly_consumption = 0
            for load in current_hub.load_set.all():
                hub_monthly_consumption += monthly_consumption_fnc(load, year, month)
            consumption_dict[month] = hub_monthly_consumption

        yearData.append(consumption_dict)

        for month in range(1, 13):
            hub_monthly_generation = 0
            for panel in current_hub.panel_set.all():
                hub_monthly_generation += monthly_generation_fnc(panel, year, month)
            generation_dict[month] = hub_monthly_generation

        yearData.append(generation_dict)

    return JsonResponse(yearData, safe=False)

def hubMonthData(request, hub_name):
    monthData = []
    consumption_dict = {}
    generation_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    if request.method == 'POST':
        month = request.POST.get('month')
        year = request.POST.get('year-month')
        month = int(month)
        year = int(year)
        day = 1
        while True:
            try:
                date(year, month, day)
                hub_daily_consumption = 0
                for load in current_hub.load_set.all():
                    hub_daily_consumption += daily_consumption_fnc(load, year, month, day)
                consumption_dict[day] = hub_daily_consumption
                day += 1
            except ValueError:
                break

        monthData.append(consumption_dict)

        day = 1
        while True:
            try:
                date(year, month, day)
                hub_daily_generation = 0
                for panel in current_hub.panel_set.all():
                    hub_daily_generation += daily_generation_fnc(panel, year, month, day)
                generation_dict[day] = hub_daily_generation
                day += 1
            except ValueError:
                break

        monthData.append(generation_dict)

    return JsonResponse(monthData, safe=False)


def hubDayData(request, hub_name):
    dayData = []
    consumption_dict = {}
    generation_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    if request.method == 'POST':
        current_date = request.POST.get('day')
        current_date = datetime.strptime(current_date, '%Y-%m-%d')

        for load in current_hub.load_set.all():
            for current_load_status in load.loadstatus_set.all():
                if (current_load_status.date_time.year == current_date.year) and (current_load_status.date_time.month == current_date.month) and (current_load_status.date_time.day == current_date.day):
                    if current_load_status.date_time.time().strftime("%H:%M") in consumption_dict:
                        consumption_dict[current_load_status.date_time.time().strftime("%H:%M")] += round(current_load_status.consumption, 2)
                    else:
                        consumption_dict[current_load_status.date_time.time().strftime("%H:%M")] = round(current_load_status.consumption, 2)

        dayData.append(consumption_dict)

        for panel in current_hub.panel_set.all():
            for current_panel_status in panel.panelstatus_set.all():
                if (current_panel_status.date_time.year == current_date.year) and (current_panel_status.date_time.month == current_date.month) and (current_panel_status.date_time.day == current_date.day):
                    if current_panel_status.date_time.time().strftime("%H:%M") in generation_dict:
                        generation_dict[current_panel_status.date_time.time().strftime("%H:%M")] += round(current_panel_status.generation, 2)
                    else:
                        generation_dict[current_panel_status.date_time.time().strftime("%H:%M")] = round(current_panel_status.generation, 2)

        dayData.append(generation_dict)

    return JsonResponse(dayData, safe=False)


def loadDetail(request, hub_name, load_name):
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    current_load = current_hub.load_set.get(load_name=load_name)
    now = datetime.now()
    max_date = date.today()
    min_date = date(max_date.year - 1, 1, 1)
    current_year = now.year
    context = {'hub': current_hub, 'load': current_load, 'years': range(current_year,current_year-2,-1), 'months': range(1,13), 'min_date': min_date, 'max_date': max_date}
    return render(request, 'PSS/load_detail.html', context)

def panelDetail(request, hub_name, panel_name):
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    current_panel = current_hub.panel_set.get(panel_name=panel_name)
    now = datetime.now()
    max_date = date.today()
    min_date = date(max_date.year - 1, 1, 1)
    current_year = now.year
    context = {'hub': current_hub, 'panel': current_panel, 'years': range(current_year,current_year-2,-1), 'months': range(1,13), 'min_date': min_date, 'max_date': max_date}
    return render(request, 'PSS/panel_detail.html', context)


def loadYearData(request, hub_name, load_name):
    yearData = []
    consumption_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    current_load = current_hub.load_set.get(load_name = load_name)
    if request.method == 'POST':
        year = request.POST.get('year')
        year = int(year)
        for month in range(1,13):
            consumption_dict[month] = monthly_consumption_fnc(current_load, year, month)

        yearData.append(consumption_dict)
    return JsonResponse(yearData, safe=False)


def loadMonthData(request, hub_name, load_name):
    monthData = []
    consumption_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    current_load = current_hub.load_set.get(load_name = load_name)
    if request.method == 'POST':
        month = request.POST.get('month')
        year = request.POST.get('year-month')
        month = int(month)
        year = int(year)
        day = 1
        while True:
            try:
                date(year, month, day)
                consumption_dict[day] = daily_consumption_fnc(current_load, year, month, day)
                day += 1
            except ValueError:
                break
        monthData.append(consumption_dict)
    return JsonResponse(monthData, safe=False)


def loadDayData(request, hub_name, load_name):
    dayData = []
    consumption_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    current_load = current_hub.load_set.get(load_name = load_name)
    if request.method == 'POST':
        current_date = request.POST.get('day')
        current_date = datetime.strptime(current_date, '%Y-%m-%d')
        for current_load_status in current_load.loadstatus_set.all():
            if (current_load_status.date_time.year == current_date.year) and (current_load_status.date_time.month == current_date.month) and (current_load_status.date_time.day == current_date.day):
                consumption_dict[current_load_status.date_time.time().strftime("%H:%M")] = round(current_load_status.consumption, 2)

        dayData.append(consumption_dict)

    return JsonResponse(dayData, safe=False)


def panelYearData(request, hub_name, panel_name):
    yearData = []
    generation_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    current_panel = current_hub.panel_set.get(panel_name = panel_name)
    if request.method == 'POST':
        year = request.POST.get('year')
        year = int(year)
        for month in range(1,13):
            generation_dict[month] = monthly_generation_fnc(current_panel, year, month)

        yearData.append(generation_dict)
    return JsonResponse(yearData, safe=False)


def panelMonthData(request, hub_name, panel_name):
    monthData = []
    generation_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    current_panel = current_hub.panel_set.get(panel_name = panel_name)
    if request.method == 'POST':
        month = request.POST.get('month')
        year = request.POST.get('year-month')
        month = int(month)
        year = int(year)
        day = 1
        while True:
            try:
                date(year, month, day)
                generation_dict[day] = daily_generation_fnc(current_panel, year, month, day)
                day += 1
            except ValueError:
                break
        monthData.append(generation_dict)
    return JsonResponse(monthData, safe=False)


def panelDayData(request, hub_name, panel_name):
    dayData = []
    generation_dict = {}
    current_user = User.objects.get(username=request.user.username)
    current_hub = current_user.hub_set.get(hub_name=hub_name)
    current_panel = current_hub.panel_set.get(panel_name = panel_name)
    if request.method == 'POST':
        current_date = request.POST.get('day')
        current_date = datetime.strptime(current_date, '%Y-%m-%d')
        for current_panel_status in current_panel.panelstatus_set.all():
            if (current_panel_status.date_time.year == current_date.year) and (current_panel_status.date_time.month == current_date.month) and (current_panel_status.date_time.day == current_date.day):
                generation_dict[current_panel_status.date_time.time().strftime("%H:%M")] = round(current_panel_status.generation, 2)
        dayData.append(generation_dict)

    return JsonResponse(dayData, safe=False)

def dataUpload(request):
    default_password = "ttbigbang"
    today = date.today()
    start_date = date(today.year, today.month, 1)
    day_count = (today - start_date).days

    if request.method == 'POST':
        data_file = request.FILES["dataFile"]
        wb = openpyxl.load_workbook(data_file)
        sheets = wb.sheetnames
        for sheet in sheets:
            worksheet = wb[sheet]
            current_user, created = User.objects.get_or_create(username=sheet)
            current_user.set_password(default_password)
            current_user.save()

            current_hub = Hub(hub_name = "hub1", owner = current_user)
            current_load = Load(load_name = "main load", hub = current_hub)
            current_panel = Panel(panel_name="main panel", hub=current_hub)
            current_hub.save()
            current_load.save()
            current_panel.save()
            for consid_date in (start_date + timedelta(days=m) for m in range(day_count + 1)):
                for row_numb in range(2,50):
                    current_datetime = datetime(consid_date.year, consid_date.month, consid_date.day, 0, 0, 0, 0) + timedelta(minutes=(30*(row_numb-2)))
                    current_panel_status = PanelStatus(panel = current_panel, generation = worksheet.cell(row = row_numb, column = 2).value, date_time = current_datetime)
                    current_panel_status.save()

                    current_load_status = LoadStatus(load=current_load, consumption = worksheet.cell(row = row_numb, column = 3).value, date_time = current_datetime)
                    current_load_status.save()

    context = {}
    return render(request, 'PSS/data_upload.html', context)


def usersInfo(request):
    users = User.objects.all()
    context = {'users': users}
    return render(request, 'PSS/users_info.html', context)


def chart(request):
    context = {'users':User.objects.all()}
    return render(request, 'PSS/chart.html', context)

def payment(request, amount):
    current_user = User.objects.get(username=request.user.username)
    context = {'user': current_user, 'amount':amount}
    return render(request, 'PSS/payment.html', context)

def charge(request):
    if request.method == 'POST':
        amount = int(request.POST.get('amount'))
        customer = stripe.Customer.create(
            name=request.POST.get('username'),
            source = request.POST['stripeToken']
        )
        charge = stripe.Charge.create(
            customer = customer,
            amount = amount*100,
            currency = 'aud',
            description = "Energy Fee"
        )
    return redirect('PSS:payment-success', amount = amount)

def paymentSuccess(request, amount):
    context = {'amount': amount}
    return render(request, 'PSS/payment_success.html', context)